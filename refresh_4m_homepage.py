#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
부품 4M 변경관리 현황 — 주간 자동 최신화 파이프라인 (소스 파일)
=================================================================

하나의 입력 엑셀(팀즈 "999. 백업" 폴더의 최신 Supplier Parts Change
Management_v0_YYMMDD.xlsx)을 받아서 아래를 자동으로 생성한다.

  1) Summary 시트를 새로 구성한 엑셀(.xlsx)
     - 담당자(ChM) × 진행 단계 COUNTIFS
     - 심의회 / QA 인정 시험 / GQMS 접수 단계는 정기·비정기 값별 분해 열 추가
     - Risk 는 4M Issue 보고의 공동 담당 셀까지 집계되도록 부분일치(COUNTIF "*이름*")
  2) 단일 파일 인터랙티브 홈페이지(.html)
     - 메뉴: 홈(진행중) / Summary(전체) / List / 4M Issue 보고
     - 홈 화면은 4M 완료·Reject·4M 완료(C)·합계 열을 제외한 진행중 중심 요약
     - 상단에서 HTML 과 원본 엑셀을 모두 다운로드(엑셀은 base64 로 내장)

데이터 출처 (두 곳)
------------------
  - List / Summary : 팀즈 '999. 백업' 폴더의 최신 'Supplier Parts Change Management_v0_YYMMDD.xlsx'
  - 4M Issue 보고  : 팀즈 '04. Meeting' 폴더의 'VS협력사변경관리팀 업무 관리_v0.xlsx' 의 '4M_Issue_보고' 시트
    (백업 파일에는 4M Issue 시트가 없으므로 별도 파일에서 읽어 생성 엑셀에 포함시킨다.)

사용법
------
  [A] 팀즈에서 직접 다운로드 (코드가 최신 파일을 내려받아 실행)
      # Microsoft Graph 액세스 토큰 필요 (환경변수 GRAPH_TOKEN)
      GRAPH_TOKEN="<access_token>" python refresh_4m_homepage.py --download [--outdir output]
      → 위 두 링크에서 최신 List 파일과 업무 관리(4M) 파일을 outdir 로 내려받고,
        그 최신 파일 기준으로 HTML·엑셀을 생성한다. 받은 원본 파일도 outdir 에 함께 저장된다.

  [B] 로컬 파일로 실행 (이미 받아둔 파일 사용)
      python refresh_4m_homepage.py "<백업최신.xlsx>" \
          --m4-file "<VS협력사변경관리팀 업무 관리_v0.xlsx>" [--outdir output]

산출물(outdir): ① 최신 List 원본  ② 최신 4M(업무관리) 원본  ③ Summary_정기비정기.xlsx  ④ 홈페이지 HTML

인증 안내
--------
  표준 Python 으로 SharePoint 에 접속하려면 인증(Graph 토큰)이 반드시 필요하다.
  · 토큰을 직접 발급/주입할 수 있는 환경: --download 로 코드가 스스로 다운로드한다.
  · 이 Cowork 플랫폼: 매주 금요일 예약 작업이 토큰 없이 자동으로 최신 파일을 받아 이 코드를 실행한다.
  · 토큰이 없으면 --download 는 안내 메시지를 출력하고 종료한다. ([B] 로컬 모드로 대체 가능)

※ 정적 HTML 자체는 보안 인증이 걸린 SharePoint 에 스스로 접속/갱신할 수 없으므로,
  "주간 자동 최신화"는 이 스크립트를 주기 실행하는 예약 작업으로 구현한다.
"""

import sys, os, re, json, base64, html, argparse, datetime, subprocess, glob
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 진행 단계 표준 순서 / 정기·비정기 분해를 붙일 단계(원본 C·E·G)
CANON_STAGES = ['4M 완료', '심의회', 'Reject', 'QA 인정 시험', '4M 완료(C)', 'GQMS 접수']
BREAKDOWN_STAGES = ['심의회', 'QA 인정 시험', 'GQMS 접수']   # 홈 화면에 노출 + 정기/비정기 분해
# 정기/비정기 값별 그래프 색상 (요구사항 지정)
PERIOD_COLORS = {
    '23년 비정기': '#E53935',       # Red
    '24년 비정기': '#E53935',       # Red
    '25년 상반기 정기': '#FB8C00',  # Orange
    '25년 하반기 정기': '#FDD835',  # Yellow
    '25년 비정기': '#FB8C00',       # Orange
    '26년 비정기': '#43A047',       # Green
    '26년 상반기 정기': '#43A047',  # Green
}
PERIOD_COLOR_DEFAULT = '#9aa0a6'
# 홈 화면에서 제외할 단독 열: 분해 없는 단계 + 합계(=요청: B·K·T·AC)
# List 화면에 노출할 핵심 컬럼(전체 컬럼은 내장 엑셀에서 확인)
LIST_KEEP = ['ChM','Req. No','정기/비정기','요청 주체','진행 단계','OEM','Project','Supplier','Part No',
 'Item Desc','CMDT','변경점','변경 사유','4M 등급','Risk','이슈','해결방안','비고','GQMS 접수',
 '1차 심의회','4M 완료 목표일','4M 완료일','진행율','Status','심의회']

# ============================ 팀즈(SharePoint) 데이터 출처 ============================
# 코드를 --download 로 실행하면 아래 두 링크에서 최신 파일을 직접 내려받는다.
#   - List/Summary : '999. 백업' 폴더 → 최신 'Supplier Parts Change Management_v0_YYMMDD.xlsx'
#   - 4M Issue 보고 : '04. Meeting' 폴더 → 'VS협력사변경관리팀 업무 관리_v0.xlsx' 의 '4M_Issue_보고' 시트
BACKUP_FOLDER_SHARE_URL = ('https://lgeteams.sharepoint.com/:f:/r/sites/O365_379639/Shared%20Documents/'
    'VS%ED%98%91%EB%A0%A5%EC%82%AC%EB%B3%80%EA%B2%BD%EA%B4%80%EB%A6%AC%ED%8C%80/'
    '06.%20Parts%20Change%20Management%20(4M)/%EB%B6%80%ED%92%88%204M%20%ED%98%84%ED%99%A9/'
    '999.%20%EB%B0%B1%EC%97%85')
M4_FILE_SHARE_URL = ('https://lgeteams.sharepoint.com/:x:/r/sites/O365_379639/Shared%20Documents/'
    'VS%ED%98%91%EB%A0%A5%EC%82%AC%EB%B3%80%EA%B2%BD%EA%B4%80%EB%A6%AC%ED%8C%80/04.%20Meeting/'
    'VS%ED%98%91%EB%A0%A5%EC%82%AC%EB%B3%80%EA%B2%BD%EA%B4%80%EB%A6%AC%ED%8C%80%20%EC%97%85%EB%AC%B4%20%EA%B4%80%EB%A6%AC_v0.xlsx')
GRAPH = 'https://graph.microsoft.com/v1.0'


def _graph_token():
    return os.environ.get('GRAPH_TOKEN') or os.environ.get('GRAPH_ACCESS_TOKEN')

def _share_id(url):
    """공유 URL → Graph /shares/{id} 용 인코딩(u! + base64url)."""
    b = base64.urlsafe_b64encode(url.encode('utf-8')).decode('ascii').rstrip('=')
    return 'u!' + b

def _g(token, url, stream=False):
    import requests
    r = requests.get(url, headers={'Authorization': 'Bearer ' + token}, stream=stream, timeout=90)
    r.raise_for_status()
    return r

def _dl_item(token, item, dest):
    url = item.get('@microsoft.graph.downloadUrl')
    if not url:
        drive = item['parentReference']['driveId']
        url = _g(token, f"{GRAPH}/drives/{drive}/items/{item['id']}").json()['@microsoft.graph.downloadUrl']
    r = _g(token, url, stream=True)
    with open(dest, 'wb') as f:
        for chunk in r.iter_content(65536):
            if chunk: f.write(chunk)
    return dest

def download_sources(outdir, token=None):
    """팀즈 두 링크에서 최신 List 파일과 4M(업무관리) 파일을 outdir 로 내려받아 (list_path, m4_path) 반환.
    표준 Python 으로 SharePoint 에 접속하려면 Microsoft Graph 액세스 토큰이 필요하다.
    (환경변수 GRAPH_TOKEN). 이 Cowork 플랫폼에서는 예약 작업이 토큰 없이 자동 다운로드한다."""
    token = token or _graph_token()
    if not token:
        raise RuntimeError(
            'Graph 액세스 토큰이 없습니다. SharePoint 다운로드에는 인증이 필요합니다.\n'
            '  · 표준 실행: 환경변수 GRAPH_TOKEN 에 Microsoft Graph 토큰을 넣고 --download 로 실행\n'
            '  · Cowork 플랫폼: 매주 금요일 예약 작업이 자동으로 최신 파일을 받아 실행합니다.')
    # 1) 백업 폴더 → 최신 List 파일
    fr = _g(token, f"{GRAPH}/shares/{_share_id(BACKUP_FOLDER_SHARE_URL)}/driveItem").json()
    drive, folder = fr['parentReference']['driveId'], fr['id']
    kids = _g(token, f"{GRAPH}/drives/{drive}/items/{folder}/children?$top=300").json()['value']
    cand = []
    for k in kids:
        m = re.search(r'_v0_(\d{6})\.xlsx$', k.get('name', ''))
        if m and 'file' in k:
            cand.append((m.group(1), k))
    if not cand:
        raise RuntimeError('백업 폴더에서 최신 List 파일(_v0_YYMMDD.xlsx)을 찾지 못했습니다.')
    cand.sort(key=lambda x: x[0])
    latest = cand[-1][1]
    list_path = _dl_item(token, latest, os.path.join(outdir, latest['name']))
    # 2) 업무관리 파일(4M Issue 보고)
    mr = _g(token, f"{GRAPH}/shares/{_share_id(M4_FILE_SHARE_URL)}/driveItem").json()
    m4_path = _dl_item(token, mr, os.path.join(outdir, mr['name']))
    return list_path, m4_path


def autofind_local(extra_dirs=None):
    """인자 없이 실행할 때, 주변 폴더에서 최신 List 파일과 4M(업무관리) 파일을 자동 탐색.
    반환: (list_path 또는 None, m4_path 또는 None)"""
    here = os.path.dirname(os.path.abspath(__file__))
    dirs = ['.', here, os.path.join(here, 'input'), os.path.join(here, 'output'),
            'input', 'output', 'grounding/downloads']
    if extra_dirs:
        dirs = list(extra_dirs) + dirs
    seen = set(); list_cands = []; m4 = None
    for d in dirs:
        for p in glob.glob(os.path.join(d, '*.xlsx')):
            ap = os.path.abspath(p)
            if ap in seen or not os.path.isfile(p):
                continue
            seen.add(ap)
            name = os.path.basename(p)
            if '_Summary_' in name or '정기비정기' in name:    # 생성 결과물은 입력에서 제외
                continue
            m = re.search(r'_v0_(\d{6})\.xlsx$', name)
            if m:
                list_cands.append((m.group(1), p))
            if ('업무 관리' in name) or ('업무관리' in name):
                m4 = m4 or p
    list_cands.sort(key=lambda x: x[0])
    return (list_cands[-1][1] if list_cands else None), m4


# ---------------------------------------------------------------- helpers
def norm(s):
    return ('' if s is None else str(s)).replace('\n', ' ').strip()

def cellval(v):
    if v is None: return ''
    if isinstance(v, (datetime.datetime, datetime.date)): return v.strftime('%Y-%m-%d')
    if isinstance(v, float): return int(v) if v == int(v) else round(v, 4)
    return v

def period_key(p):
    """'26년 상반기 정기','25년 비정기' 등을 연도→상/하반기→비정기 순으로 정렬."""
    p = str(p)
    m = re.search(r'(\d+)\s*년', p)
    yr = int(m.group(1)) if m else 99
    half = 1 if '상반기' in p else (2 if '하반기' in p else 3)  # 상→하→비정기
    return (yr, half, p)

def find_sheet(wb, *needles):
    for ws in wb.worksheets:
        nm = ws.title.lower().replace(' ', '').replace('_', '')
        if all(n.lower().replace(' ', '').replace('_', '') in nm for n in needles):
            return ws
    return None

def col_by_header(ws, header_row, names):
    """header_row 행에서 주어진 후보명과 일치하는 1-based 컬럼 인덱스."""
    want = [norm(n) for n in names]
    for c in range(1, ws.max_column + 1):
        if norm(ws.cell(row=header_row, column=c).value) in want:
            return c
    return None

def date_from_name(path):
    m = re.search(r'_v0_.*?(\d{6})', os.path.basename(path))
    if not m: return datetime.date.today().strftime('%Y-%m-%d')
    s = m.group(1)
    try:
        return datetime.datetime.strptime(s, '%y%m%d').strftime('%Y-%m-%d')
    except ValueError:
        return s


# ---------------------------------------------------------------- analyze
def analyze(wb, m4ws=None):
    listws = wb['List'] if 'List' in wb.sheetnames else find_sheet(wb, 'list')
    if m4ws is None:
        m4ws = find_sheet(wb, '4m') or find_sheet(wb, 'issue') or find_sheet(wb, '보고')
    if listws is None:
        raise SystemExit('List 시트를 찾을 수 없습니다.')

    # List 헤더는 3행
    HROW = 3
    c_chm = col_by_header(listws, HROW, ['ChM']) or 2
    c_per = col_by_header(listws, HROW, ['정기/비정기']) or 4
    c_stg = col_by_header(listws, HROW, ['진행 단계', '진행단계']) or 6
    c_bigo = col_by_header(listws, HROW, ['비고']) or 27

    people_total = {}
    stage_set = []
    period_set = []
    cntS = {}   # (chm,stage)
    cntP = {}   # (chm,stage,period)
    cntNC = {}  # (chm) 완료불가(비고 보유)
    for r in range(HROW + 1, listws.max_row + 1):
        chm = norm(listws.cell(row=r, column=c_chm).value)
        if not chm: continue
        stg = norm(listws.cell(row=r, column=c_stg).value)
        per = norm(listws.cell(row=r, column=c_per).value)
        bigo = listws.cell(row=r, column=c_bigo).value
        people_total[chm] = people_total.get(chm, 0) + 1
        if stg and stg not in stage_set: stage_set.append(stg)
        if per and per not in period_set: period_set.append(per)
        cntS[(chm, stg)] = cntS.get((chm, stg), 0) + 1
        cntP[(chm, stg, per)] = cntP.get((chm, stg, per), 0) + 1
        if bigo not in (None, ''):
            cntNC[chm] = cntNC.get(chm, 0) + 1

    # 진행 단계 순서: 표준 우선, 그 외 발견된 것 뒤에 추가
    stages = [s for s in CANON_STAGES if s in stage_set] + [s for s in stage_set if s not in CANON_STAGES]
    periods = sorted([p for p in period_set if p], key=period_key)
    # 담당자: 합계 내림차순
    people = sorted(people_total.keys(), key=lambda p: -people_total[p])

    # Risk: 4M Issue 보고 담당자 열의 공동 담당 셀까지 부분일치 집계
    risk = {p: 0 for p in people}
    m4_owner_names = []
    if m4ws is not None:
        c_owner = col_by_header(m4ws, 2, ['담당자']) or 2
        for r in range(3, m4ws.max_row + 1):
            v = m4ws.cell(row=r, column=c_owner).value
            if v not in (None, ''):
                m4_owner_names.append(str(v))
        for p in people:
            risk[p] = sum(1 for v in m4_owner_names if p in v)

    return dict(listws=listws, m4ws=m4ws, HROW=HROW,
                c_chm=c_chm, c_per=c_per, c_stg=c_stg, c_bigo=c_bigo,
                people=people, stages=stages, periods=periods,
                cntS=cntS, cntP=cntP, cntNC=cntNC, risk=risk)


# ---------------------------------------------------------------- summary sheet (formulas)
def build_summary_sheet(wb, A):
    if 'Summary' in wb.sheetnames:
        wb.remove(wb['Summary'])
    ws = wb.create_sheet('Summary', 0)
    people, stages, periods = A['people'], A['stages'], A['periods']
    listws = A['listws']; m4name = A['m4ws'].title if A['m4ws'] else '4M_Issue_보고'
    LB = listws.title
    cl = get_column_letter
    Lchm = cl(A['c_chm']); Lper = cl(A['c_per']); Lstg = cl(A['c_stg']); Lbg = cl(A['c_bigo'])

    # 레이아웃 컬럼 배치
    ws.cell(row=2, column=1, value='ChM / 진행 단계')
    col = 2
    stage_col = {}; block_cols = {}
    for st in stages:
        stage_col[st] = cl(col); ws.cell(row=2, column=col, value=st); col += 1
        if st in BREAKDOWN_STAGES:
            bl = []
            for pv in periods:
                L = cl(col); ws.cell(row=2, column=col, value=pv); bl.append((L, pv)); col += 1
            block_cols[st] = bl
    SUM = cl(col); ws.cell(row=2, column=col, value='합계'); col += 1
    INP = cl(col); ws.cell(row=2, column=col, value='합계(진행중)'); col += 1
    NC = cl(col); ws.cell(row=2, column=col, value='완료불가사유'); col += 1
    RK = cl(col); ws.cell(row=2, column=col, value='Risk 보유현황'); col += 1
    LAST = col - 1

    all_stage_letters = [stage_col[s] for s in stages]
    inprog_letters = [stage_col[s] for s in stages if s in BREAKDOWN_STAGES]

    # 1행 제목/그룹 라벨
    ws.cell(row=1, column=1, value='ChM별 진행 단계 현황')
    for st in stages:
        if st in BREAKDOWN_STAGES:
            bl = block_cols[st]; first, last = bl[0][0], bl[-1][0]
            ws.cell(row=1, column=column_index_from_string(first), value=f'{st} (정기/비정기별)')
            ws.merge_cells(f'{first}1:{last}1')
    ws[f'{NC}1'] = '비고열에 값이 있는 경우'
    ws[f'{RK}1'] = 'Risk(4M Issue) 보유현황'

    # 데이터 행(담당자) — 함수
    for i, name in enumerate(people):
        r = 3 + i
        ws.cell(row=r, column=1, value=name)
        for st in stages:
            sc = stage_col[st]
            ws[f'{sc}{r}'] = f'=COUNTIFS({LB}!${Lchm}$4:${Lchm}$9999,$A{r},{LB}!${Lstg}$4:${Lstg}$9999,{sc}$2)'
            if st in BREAKDOWN_STAGES:
                for L, pv in block_cols[st]:
                    ws[f'{L}{r}'] = (f'=COUNTIFS({LB}!${Lchm}$4:${Lchm}$9999,$A{r},'
                                     f'{LB}!${Lstg}$4:${Lstg}$9999,{sc}$2,'
                                     f'{LB}!${Lper}$4:${Lper}$9999,{L}$2)')
        ws[f'{SUM}{r}'] = '=' + '+'.join(f'{c}{r}' for c in all_stage_letters)
        ws[f'{INP}{r}'] = '=' + '+'.join(f'{c}{r}' for c in inprog_letters)
        ws[f'{NC}{r}'] = f'=COUNTIFS({LB}!${Lchm}:${Lchm},$A{r},{LB}!${Lbg}:${Lbg},"<>")'
        ws[f'{RK}{r}'] = f'=COUNTIF(\'{m4name}\'!$B:$B,"*"&$A{r}&"*")'

    # 합계 행
    TR = 3 + len(people)
    ws.cell(row=TR, column=1, value='합계')
    for ci in range(2, LAST + 1):
        L = cl(ci); ws[f'{L}{TR}'] = f'=SUM({L}3:{L}{TR-1})'

    _style_summary(ws, stages, stage_col, block_cols, SUM, INP, NC, RK, LAST, TR)
    return dict(stage_col=stage_col, block_cols=block_cols, SUM=SUM, INP=INP, NC=NC, RK=RK,
                LAST=LAST, TR=TR)


def _style_summary(ws, stages, stage_col, block_cols, SUM, INP, NC, RK, LAST, TR):
    cl = get_column_letter
    LG = PatternFill('solid', fgColor='A50034'); LGs = PatternFill('solid', fgColor='D9536F')
    GRP = PatternFill('solid', fgColor='F3E1E7'); TOT = PatternFill('solid', fgColor='F4F5F7')
    thin = Side(style='thin', color='D9D9D9'); bd = Border(thin, thin, thin, thin)
    cen = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cenL = Alignment(horizontal='left', vertical='center')
    ws['A1'].font = Font(bold=True, size=13, color='A50034')
    for st in stages:
        if st in block_cols:
            c = ws.cell(row=1, column=column_index_from_string(block_cols[st][0][0]))
            c.font = Font(bold=True, size=10, color='A50034'); c.alignment = cen; c.fill = GRP
    for cc in (NC, RK):
        ws[f'{cc}1'].font = Font(italic=True, size=8, color='888888'); ws[f'{cc}1'].alignment = cen
    stage_letters = set(list(stage_col.values()) + [SUM, INP, NC, RK])
    for ci in range(1, LAST + 1):
        L = cl(ci); c = ws[f'{L}2']; c.alignment = cen; c.border = bd
        if L in stage_letters or ci == 1:
            c.fill = LG; c.font = Font(bold=True, color='FFFFFF', size=10)
        else:
            c.fill = LGs; c.font = Font(bold=True, color='FFFFFF', size=9)
    for r in range(3, TR + 1):
        for ci in range(1, LAST + 1):
            L = cl(ci); c = ws[f'{L}{r}']; c.border = bd
            if ci == 1:
                c.alignment = cenL; c.font = Font(bold=True, size=10, color=('A50034' if r == TR else '1D1D1F'))
            else:
                c.alignment = Alignment(horizontal='center', vertical='center'); c.font = Font(bold=(r == TR), size=10)
            if r == TR: c.fill = TOT
    ws.column_dimensions['A'].width = 13
    for ci in range(2, LAST + 1):
        L = cl(ci)
        if L in (SUM, INP, NC, RK): ws.column_dimensions[L].width = 11
        elif L in stage_col.values(): ws.column_dimensions[L].width = 10
        else: ws.column_dimensions[L].width = 12.5
    ws.row_dimensions[1].height = 20; ws.row_dimensions[2].height = 42
    ws.freeze_panes = 'B3'


def try_recalc(path):
    """LibreOffice 가 있으면 수식 값을 미리 계산해 캐시한다(안전 모드).
    재계산이 실패하거나 파일을 손상시키면 수식만 들어있는 원본으로 자동 복원하므로,
    어떤 경우에도 유효한 .xlsx 가 보장된다. (수식은 Excel 에서 열 때 자동 계산됨)"""
    import shutil, zipfile
    skill = '/opt/workspace-config/.claude/skills/xlsx/scripts/recalc.py'
    if not os.path.exists(skill):
        return False
    backup = path + '.bak'
    shutil.copy2(path, backup)
    baked = False
    try:
        subprocess.run([sys.executable, skill, path], cwd=os.path.dirname(skill),
                       capture_output=True, timeout=180)
        if zipfile.is_zipfile(path):
            ws = openpyxl.load_workbook(path, data_only=True)['Summary']
            baked = ws['AC3'].value not in (None, '')   # 값이 실제로 채워졌는지 확인
    except Exception:
        baked = False
    if not baked:
        shutil.copy2(backup, path)        # 손상/미베이킹 시 수식본 복원
    try: os.remove(backup)
    except OSError: pass
    return baked


# ---------------------------------------------------------------- HTML build
def build_html(A, xlsx_path, out_html, basis_date):
    listws = A['listws']; m4ws = A['m4ws']
    people, stages, periods = A['people'], A['stages'], A['periods']
    cntS, cntP, cntNC, risk = A['cntS'], A['cntP'], A['cntNC'], A['risk']

    # 요약 값(파이썬 직접 계산 → 엑셀 재계산과 무관하게 항상 정확)
    def stage_tot(p, s): return cntS.get((p, s), 0)
    def per_val(p, s, pv): return cntP.get((p, s, pv), 0)
    inprog = {p: sum(stage_tot(p, s) for s in stages if s in BREAKDOWN_STAGES) for p in people}
    grand = {p: sum(stage_tot(p, s) for s in stages) for p in people}

    # List 추출(핵심 컬럼)
    HROW = A['HROW']; maxc = listws.max_column
    headers = [norm(listws.cell(row=HROW, column=c).value) for c in range(1, maxc + 1)]
    keep_idx = [i for i, h in enumerate(headers) if h in LIST_KEEP]
    lh = [headers[i] for i in keep_idx]
    ld = []
    for r in range(HROW + 1, listws.max_row + 1):
        row = [cellval(listws.cell(row=r, column=i + 1).value) for i in keep_idx]
        if any(x != '' for x in row): ld.append(row)

    # 4M Issue 추출
    m4h = ['담당자', '최초보고', '업데이트', '부족시점', '부품/협력사', 'OEM/PJT',
           '변경점 및 주요 이슈사항', '상세 진행현황 및 향후 대응방안', '상태', '비고']
    m4 = []
    if m4ws is not None:
        for r in range(3, m4ws.max_row + 1):
            row = [cellval(m4ws.cell(row=r, column=c).value) for c in range(2, 12)]
            if str(row[0]).strip() or (len(row) > 6 and str(row[6]).strip()):
                m4.append(row)

    xb64 = base64.b64encode(open(xlsx_path, 'rb').read()).decode()
    xname = os.path.basename(xlsx_path)

    def num(v): return f'{int(v):,}' if isinstance(v, (int,)) else html.escape(str(v))

    # ---- summary table renderer (home=False excludes 분해없는 단계 + 합계) ----
    def render_summary(full):
        def rowvals(p):
            d = {'A': p}
            for s in stages:
                d[s] = stage_tot(p, s)
                if s in BREAKDOWN_STAGES:
                    for pv in periods: d[(s, pv)] = per_val(p, s, pv)
            d['합계'] = grand[p]; d['진행중'] = inprog[p]; d['NC'] = cntNC.get(p, 0); d['RK'] = risk.get(p, 0)
            return d
        rows = [rowvals(p) for p in people]
        tot = {'A': '합계'}
        for s in stages:
            tot[s] = sum(stage_tot(p, s) for p in people)
            if s in BREAKDOWN_STAGES:
                for pv in periods: tot[(s, pv)] = sum(per_val(p, s, pv) for p in people)
        tot['합계'] = sum(grand.values()); tot['진행중'] = sum(inprog.values())
        tot['NC'] = sum(cntNC.values()); tot['RK'] = sum(risk.values())

        layout = [('single', 'A', '담당자')]
        for s in stages:
            if s in BREAKDOWN_STAGES:
                layout.append(('group', s))
            elif full:
                layout.append(('single', s, s))
        if full: layout.append(('single', '합계', '합계'))
        layout += [('single', '진행중', '합계(진행중)'), ('single', 'NC', '완료불가사유'), ('single', 'RK', 'Risk 보유현황')]

        r1, r2 = [], []
        for it in layout:
            if it[0] == 'single':
                r1.append(f'<th rowspan="2" class="sng">{html.escape(it[2])}</th>')
            else:
                s = it[1]
                r1.append(f'<th colspan="{1+len(periods)}" class="grp">{html.escape(s)}</th>')
                r2.append('<th class="tcol">계</th>')
                for pv in periods: r2.append(f'<th class="sub">{html.escape(pv)}</th>')
        thead = f'<tr>{"".join(r1)}</tr><tr>{"".join(r2)}</tr>'
        body = ''
        for row in rows + [tot]:
            cls = ' class="tot"' if row is tot else ''
            cells = ''
            for it in layout:
                if it[0] == 'single':
                    k = it[1]
                    if k == 'A': cells += f'<th class="rowname">{html.escape(str(row["A"]))}</th>'
                    else: cells += f'<td>{num(row[k])}</td>'
                else:
                    s = it[1]; cells += f'<td class="tcol">{num(row[s])}</td>'
                    for pv in periods: cells += f'<td class="subv">{num(row[(s,pv)])}</td>'
            body += f'<tr{cls}>{cells}</tr>'
        return f'<div class="tablewrap"><table class="sumtbl"><thead>{thead}</thead><tbody>{body}</tbody></table></div>', tot

    HOME_TBL, tot = render_summary(False)
    FULL_TBL, _ = render_summary(True)

    KPI = [('합계(진행중)', tot['진행중'], '심의회+QA+GQMS', 'var(--lg)')]
    for s in BREAKDOWN_STAGES:
        if s in stages:
            KPI.append((s, tot.get(s, 0), 'in-progress', 'var(--warn)' if s != '심의회' else 'var(--lg)'))
    KPI += [('완료불가사유', tot['NC'], '비고열 보유', 'var(--muted)'),
            ('Risk 보유현황', tot['RK'], '4M Issue 등록', 'var(--bad)')]
    kpihtml = ''.join(f'<div class="kpi"><div class="lab"><span class="dot" style="background:{c}"></span>'
                      f'{html.escape(l)}</div><div class="num">{v:,}</div><div class="ft">{html.escape(f)}</div></div>'
                      for l, v, f, c in KPI)

    chart = sorted([{'nm': p, 'sim': stage_tot(p, '심의회'), 'qa': stage_tot(p, 'QA 인정 시험'),
                     'gq': stage_tot(p, 'GQMS 접수'), 'tot': inprog[p]} for p in people],
                   key=lambda x: -x['tot'])
    CHART_JSON = json.dumps(chart, ensure_ascii=False)

    # 정기/비정기 상태별 진행중 건수 (심의회+QA+GQMS 합산) — periods 순서 유지
    inprog_stages = [s for s in stages if s in BREAKDOWN_STAGES]
    period_chart = [{'label': pv,
                     'value': sum(per_val(p, s, pv) for p in people for s in inprog_stages),
                     'color': PERIOD_COLORS.get(pv, PERIOD_COLOR_DEFAULT)} for pv in periods]
    PERIOD_JSON = json.dumps(period_chart, ensure_ascii=False)
    LISTDATA = json.dumps({'listHeader': lh, 'list': ld, 'm4Header': m4h, 'm4': m4}, ensure_ascii=False)

    DL = (f'<a class="dl" download="{html.escape(xname)}" '
          f'href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{xb64}">⬇ 원본 엑셀</a>')

    page = HTML_TMPL.format(CSS=CSS, basis=html.escape(basis_date), DL=DL, kpihtml=kpihtml,
                            HOME_TBL=HOME_TBL, FULL_TBL=FULL_TBL,
                            LISTDATA=LISTDATA, CHART_JSON=CHART_JSON, PERIOD_JSON=PERIOD_JSON, JS=JS)
    open(out_html, 'w', encoding='utf-8').write(page)
    return out_html


# ---- 정적 CSS / JS / HTML 템플릿 ----
CSS = r"""
:root{--lg:#A50034;--lg2:#C00043;--ink:#1d1d1f;--muted:#6b6f76;--line:#e6e7eb;--bg:#f4f5f7;
--card:#fff;--soft:#faf7f8;--ok:#1a7f4b;--warn:#b9770a;--bad:#c0392b;--sim:#A50034;--qa:#E08A1E;--gq:#2E7D9A;
--shadow:0 1px 3px rgba(0,0,0,.06),0 8px 24px rgba(0,0,0,.05);}
@media (prefers-color-scheme: dark){:root{--ink:#e9eaee;--muted:#a0a4ad;--line:#2b2d33;--bg:#15161a;--card:#1e2026;--soft:#23252c;}}
*{box-sizing:border-box}html,body{margin:0;padding:0}
body{font-family:"Segoe UI","Malgun Gothic","Apple SD Gothic Neo",system-ui,sans-serif;color:var(--ink);background:var(--bg);line-height:1.5}
header.top{position:sticky;top:0;z-index:50;background:linear-gradient(100deg,var(--lg),var(--lg2));color:#fff;box-shadow:0 2px 14px rgba(165,0,52,.28)}
.bar{max-width:1400px;margin:0 auto;display:flex;align-items:center;gap:10px;padding:0 20px;height:60px;flex-wrap:wrap}
.brand{display:flex;flex-direction:column;line-height:1.1;margin-inline-end:auto}
.brand b{font-size:17px}.brand span{font-size:11.5px;opacity:.86}
nav.menu{display:flex;gap:4px}
nav.menu button{appearance:none;border:0;background:transparent;color:#fff;cursor:pointer;font-size:14px;font-weight:600;padding:9px 15px;border-radius:999px;opacity:.82;transition:.15s}
nav.menu button:hover{opacity:1;background:rgba(255,255,255,.14)}
nav.menu button.active{opacity:1;background:#fff;color:var(--lg)}
.dls{display:flex;gap:7px}
.dl{display:inline-flex;align-items:center;gap:6px;font-size:12px;font-weight:700;text-decoration:none;cursor:pointer;border:1px solid rgba(255,255,255,.55);color:#fff;background:rgba(255,255,255,.12);padding:7px 11px;border-radius:9px}
.dl:hover{background:#fff;color:var(--lg)}
main{max-width:1400px;margin:0 auto;padding:24px 20px 60px}
.page{display:none}.page.show{display:block;animation:fade .25s}
@keyframes fade{from{opacity:0;transform:translateY(6px)}to{opacity:1;transform:none}}
h1.pt{font-size:22px;margin:4px 0 2px}.sub{color:var(--muted);font-size:13px;margin-bottom:20px}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(168px,1fr));gap:13px;margin-bottom:22px}
.kpi{background:var(--card);border:1px solid var(--line);border-radius:16px;padding:16px 16px 14px;box-shadow:var(--shadow)}
.kpi .lab{font-size:12px;color:var(--muted);font-weight:600;display:flex;align-items:center;gap:7px}
.kpi .dot{width:9px;height:9px;border-radius:3px}
.kpi .num{font-size:27px;font-weight:800;margin-top:6px;letter-spacing:-.5px}
.kpi .ft{font-size:11px;color:var(--muted);margin-top:2px}
.card{background:var(--card);border:1px solid var(--line);border-radius:16px;box-shadow:var(--shadow);overflow:hidden;margin-bottom:18px}
.card h3{margin:0;padding:14px 18px;font-size:14.5px;border-bottom:1px solid var(--line);display:flex;align-items:center;gap:8px}
.card h3 .b3{width:4px;height:15px;background:var(--lg);border-radius:2px}
.tablewrap{overflow:auto;max-height:74vh;border:1px solid var(--line);border-radius:14px;background:var(--card);box-shadow:var(--shadow)}
.card .tablewrap{border:0;border-radius:0;box-shadow:none}
table{border-collapse:collapse;width:100%;font-size:12.5px}
th,td{padding:8px 11px;border-bottom:1px solid var(--line);white-space:nowrap;text-align:center}
thead th{position:sticky;top:0;background:var(--soft);z-index:3}
.sumtbl th.grp{background:var(--lg);color:#fff;font-size:11.5px;border-inline:1px solid #fff}
.sumtbl th.sng{background:#7a0026;color:#fff;font-size:11px;vertical-align:middle}
.sumtbl thead tr:nth-child(2) th{top:33px}
.sumtbl th.tcol{color:var(--lg);font-weight:800;font-size:11px}
.sumtbl th.sub{background:var(--soft);color:var(--muted);font-size:10.5px;font-weight:700;min-width:60px;white-space:normal;line-height:1.25}
.sumtbl td{font-variant-numeric:tabular-nums}
.sumtbl td.tcol{font-weight:800;background:rgba(165,0,52,.05);color:var(--lg)}
.sumtbl .rowname{position:sticky;left:0;background:var(--card);text-align:left;font-weight:700;z-index:2;border-inline-end:2px solid var(--line)}
.sumtbl tr.tot td,.sumtbl tr.tot .rowname{background:var(--soft);font-weight:800}
.sumtbl tr.tot td.tcol{background:rgba(165,0,52,.1)}
tbody tr:hover td{background:rgba(165,0,52,.03)}
tbody tr:hover .rowname{background:var(--soft)}
.chart{padding:10px 18px 16px}
.brow{display:grid;grid-template-columns:64px 1fr 190px;align-items:center;gap:10px;margin:8px 0}
.brow .nm{font-size:12.5px;font-weight:600}
.stack{display:flex;height:20px;border-radius:7px;overflow:hidden;background:rgba(0,0,0,.05)}
.stack i{display:block;height:100%}
.brow .vl{font-size:11px;color:var(--muted);text-align:right;font-variant-numeric:tabular-nums}
.legend{display:flex;gap:16px;padding:4px 18px 0;font-size:12px;color:var(--muted);flex-wrap:wrap}
.legend span{display:inline-flex;align-items:center;gap:6px}.legend i{width:11px;height:11px;border-radius:3px;display:inline-block}
.pchart{padding:12px 18px 18px}
.prow{display:grid;grid-template-columns:120px 1fr 96px;align-items:center;gap:12px;margin:9px 0}
.prow .pnm{font-size:12.5px;font-weight:700;white-space:nowrap}
.ptrack{background:rgba(0,0,0,.05);border-radius:8px;height:22px;overflow:hidden}
.pfill{height:100%;border-radius:8px;min-width:2px;cursor:pointer;transition:filter .12s}
.pfill:hover{filter:brightness(1.08) saturate(1.1)}
.prow .pvl{font-size:12px;text-align:right;font-variant-numeric:tabular-nums;color:var(--muted);font-weight:700}
.tip{position:fixed;pointer-events:none;background:#1d1d1f;color:#fff;padding:6px 10px;border-radius:8px;
font-size:12px;font-weight:600;line-height:1.35;z-index:999;box-shadow:0 4px 16px rgba(0,0,0,.25);max-width:240px}
.tip b{color:#ffd2dd}
.toolbar{display:flex;flex-wrap:wrap;gap:10px;align-items:center;margin-bottom:14px}
.inp{flex:1;min-width:200px;display:flex;align-items:center;gap:8px;background:var(--card);border:1px solid var(--line);border-radius:11px;padding:9px 13px;box-shadow:var(--shadow)}
.inp input{border:0;outline:0;background:transparent;color:var(--ink);font-size:13.5px;width:100%}
.inp svg{flex:none;opacity:.5}
select{background:var(--card);border:1px solid var(--line);border-radius:11px;padding:9px 13px;font-size:13px;color:var(--ink);cursor:pointer;box-shadow:var(--shadow);font-weight:600}
.count{font-size:12.5px;color:var(--muted);font-weight:600}
.listwrap td{max-width:300px;white-space:normal;word-break:break-word;text-align:left}
.listwrap th{cursor:pointer;text-align:left}
.pager{display:flex;gap:8px;align-items:center;justify-content:flex-end;margin-top:14px;flex-wrap:wrap}
.pager button{padding:7px 12px;border-radius:9px;border:1px solid var(--line);background:var(--card);cursor:pointer;font-weight:700;font-size:13px;color:var(--ink)}
.pager button:disabled{opacity:.4;cursor:default}.pill{font-size:11px;color:var(--muted);font-weight:600}
.badge{display:inline-block;padding:2px 9px;border-radius:999px;font-size:11px;font-weight:800}
.b-open{background:rgba(192,57,43,.12);color:var(--bad)}.b-closed{background:rgba(26,127,75,.13);color:var(--ok)}.b-grey{background:rgba(120,120,120,.14);color:var(--muted)}
.cards4m{display:grid;grid-template-columns:repeat(auto-fill,minmax(420px,1fr));gap:16px}
@media(max-width:720px){.cards4m{grid-template-columns:1fr}}
.issue{background:var(--card);border:1px solid var(--line);border-radius:16px;box-shadow:var(--shadow);overflow:hidden;border-top:3px solid var(--lg)}
.issue .ihd{padding:13px 16px 11px;border-bottom:1px solid var(--line);background:var(--soft)}
.issue .t1{font-size:15px;font-weight:800;display:flex;align-items:center;gap:9px;flex-wrap:wrap}
.issue .meta{display:flex;flex-wrap:wrap;gap:6px 14px;margin-top:8px;font-size:11.7px;color:var(--muted)}
.issue .meta b{color:var(--ink);font-weight:700}
.issue .ibd{padding:13px 16px}.fld{margin-bottom:12px}.fld:last-child{margin-bottom:0}
.fld .fl{font-size:11px;font-weight:800;color:var(--lg);margin-bottom:4px}
.fld .fv{font-size:12.6px;white-space:pre-wrap;word-break:break-word}
.empty{padding:40px;text-align:center;color:var(--muted)}
.note{font-size:12px;color:var(--muted);margin:10px 2px 0}
footer{max-width:1400px;margin:0 auto;padding:8px 20px 40px;color:var(--muted);font-size:11.5px}
"""

JS = r"""
const RAW=JSON.parse(document.getElementById('ldata').textContent);
const CHART=JSON.parse(document.getElementById('cdata').textContent);
const $=(s,e=document)=>e.querySelector(s),$$=(s,e=document)=>[...e.querySelectorAll(s)];
function esc(s){return (s==null?'':String(s)).replace(/[&<>]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[c]));}
// 마우스 호버 툴팁(수치 표시)
const TIP=document.createElement('div');TIP.className='tip';TIP.style.display='none';document.body.appendChild(TIP);
function bindTips(){$$('.hastip').forEach(el=>{ if(el._tb)return; el._tb=1;
  el.addEventListener('mouseenter',()=>{TIP.innerHTML=el.dataset.tip;TIP.style.display='block';});
  el.addEventListener('mousemove',e=>{TIP.style.left=Math.min(e.clientX+14,window.innerWidth-250)+'px';TIP.style.top=(e.clientY+16)+'px';});
  el.addEventListener('mouseleave',()=>{TIP.style.display='none';});});}
$$('nav.menu button').forEach(b=>b.addEventListener('click',()=>{
  $$('nav.menu button').forEach(x=>x.classList.remove('active'));b.classList.add('active');
  $$('.page').forEach(p=>p.classList.remove('show'));$('#'+b.dataset.t).classList.add('show');
  window.scrollTo({top:0,behavior:'smooth'});}));
$('#dlhtml').addEventListener('click',()=>{
  const h='<!doctype html>\n'+document.documentElement.outerHTML;
  const bl=new Blob([h],{type:'text/html;charset=utf-8'});const a=document.createElement('a');
  a.href=URL.createObjectURL(bl);a.download='부품4M_변경관리_현황.html';document.body.appendChild(a);a.click();
  a.remove();setTimeout(()=>URL.revokeObjectURL(a.href),1500);});
// 정기/비정기 상태별 진행중 그래프 (지정 색상 + 호버 수치)
(function(){const P=JSON.parse(document.getElementById('pdata').textContent);
 const tot=P.reduce((s,d)=>s+d.value,0)||1;const mx=Math.max(1,...P.map(d=>d.value));
 $('#pchart').innerHTML=P.map(d=>{const pct=(d.value/tot*100);
  return `<div class="prow"><div class="pnm">${esc(d.label)}</div>
   <div class="ptrack"><div class="pfill hastip" style="width:${(d.value/mx*100).toFixed(1)}%;background:${d.color}" data-tip="<b>${esc(d.label)}</b><br>${d.value.toLocaleString()}건 · 진행중의 ${pct.toFixed(1)}%"></div></div>
   <div class="pvl">${d.value.toLocaleString()}건</div></div>`;}).join('');
 bindTips();})();
// 담당자별 진행중 구성 (세그먼트 호버 수치)
(function(){const mx=Math.max(1,...CHART.map(c=>c.tot));
 $('#chart').innerHTML=CHART.map(c=>{const w=c.tot/mx*100;const t=c.tot||1;
  const seg=(v,col,nm)=>v?`<i class="hastip" style="width:${v/t*100}%;background:${col}" data-tip="<b>${esc(c.nm)} · ${nm}</b><br>${v.toLocaleString()}건"></i>`:'';
  return `<div class="brow"><div class="nm">${esc(c.nm)}</div>
   <div class="stack" style="width:${w}%;min-width:2px">${seg(c.sim,'var(--sim)','심의회')}${seg(c.qa,'var(--qa)','QA 인정 시험')}${seg(c.gq,'var(--gq)','GQMS 접수')}</div>
   <div class="vl">계 ${c.tot.toLocaleString()} (심${c.sim}/QA${c.qa}/GQ${c.gq})</div></div>`;}).join('');
 bindTips();})();
(function(){let page=1,per=50,q='',si=-1,sa=true;
 const head=$('#lhead'),body=$('#lbody'),pg=$('#lpager'),H=RAW.listHeader;
 function rows(){let r=RAW.list;if(q){const t=q.toLowerCase();r=r.filter(x=>x.some(c=>String(c).toLowerCase().includes(t)));}
  if(si>=0){r=[...r].sort((a,b)=>{let x=a[si],y=b[si];const nx=parseFloat(x),ny=parseFloat(y);
   if(!isNaN(nx)&&!isNaN(ny))return sa?nx-ny:ny-nx;return sa?String(x).localeCompare(String(y),'ko'):String(y).localeCompare(String(x),'ko');});}return r;}
 function render(){const all=rows(),pages=Math.max(1,Math.ceil(all.length/per));if(page>pages)page=pages;
  const sl=all.slice((page-1)*per,page*per);
  head.innerHTML='<tr>'+H.map((h,i)=>`<th data-i="${i}">${esc(h)}${si===i?(sa?' ▲':' ▼'):''}</th>`).join('')+'</tr>';
  body.innerHTML=sl.map(r=>'<tr>'+r.map(c=>`<td>${esc(c)}</td>`).join('')+'</tr>').join('')||`<tr><td colspan="${H.length}" class="empty">검색 결과가 없습니다</td></tr>`;
  pg.innerHTML=`<span class="count">총 ${all.length.toLocaleString()}건</span>
   <button data-p="1" ${page<=1?'disabled':''}>«</button><button data-p="${page-1}" ${page<=1?'disabled':''}>이전</button>
   <span class="pill">${page} / ${pages}</span><button data-p="${page+1}" ${page>=pages?'disabled':''}>다음</button><button data-p="${pages}" ${page>=pages?'disabled':''}>»</button>`;
  $$('button',pg).forEach(b=>b.addEventListener('click',()=>{page=+b.dataset.p;render();}));
  $$('th',head).forEach(th=>th.addEventListener('click',()=>{const i=+th.dataset.i;if(si===i)sa=!sa;else{si=i;sa=true;}render();}));}
 $('#lsearch').addEventListener('input',e=>{q=e.target.value;page=1;render();});
 $('#lper').addEventListener('change',e=>{per=+e.target.value;page=1;render();});render();})();
(function(){const H=RAW.m4Header,rows=RAW.m4,I={};H.forEach((h,i)=>I[h]=i);
 const wrap=$('#m4cards'),cnt=$('#m4count'),selS=$('#m4status');
 const st=[...new Set(rows.map(r=>String(r[I['상태']]||'').trim()).filter(Boolean))];
 selS.innerHTML='<option value="">상태 전체</option>'+st.map(s=>`<option>${esc(s)}</option>`).join('');
 let q='',fs='';
 const badge=s=>{const t=String(s||'').toLowerCase();return t==='open'?'<span class="badge b-open">Open</span>':t==='closed'?'<span class="badge b-closed">Closed</span>':s?`<span class="badge b-grey">${esc(s)}</span>`:'';};
 const fld=(l,v)=>{v=(v==null?'':String(v)).trim();return v?`<div class="fld"><div class="fl">${esc(l)}</div><div class="fv">${esc(v)}</div></div>`:'';};
 function render(){let r=rows.filter(row=>{if(fs&&String(row[I['상태']]||'').trim()!==fs)return false;
   if(q){const t=q.toLowerCase();return row.some(c=>String(c).toLowerCase().includes(t));}return true;});
  cnt.textContent=`${r.length}건`;
  wrap.innerHTML=r.map(row=>{const g=k=>row[I[k]];return `<div class="issue"><div class="ihd">
   <div class="t1">${esc(String(g('부품/협력사')||'').replace(/\n/g,' '))} ${badge(g('상태'))}</div>
   <div class="meta"><span>담당자 <b>${esc(String(g('담당자')||'').replace(/\n/g,', '))}</b></span>
   <span>최초보고 <b>${esc(g('최초보고')||'-')}</b></span><span>업데이트 <b>${esc(g('업데이트')||'-')}</b></span>
   <span>부족시점 <b>${esc(g('부족시점')||'-')}</b></span></div></div><div class="ibd">
   ${fld('OEM / PJT',g('OEM/PJT'))}${fld('변경점 및 주요 이슈사항',g('변경점 및 주요 이슈사항'))}
   ${fld('상세 진행현황 및 향후 대응방안',g('상세 진행현황 및 향후 대응방안'))}${fld('비고',g('비고'))}</div></div>`;}).join('')||'<div class="empty">결과가 없습니다</div>';}
 selS.addEventListener('change',e=>{fs=e.target.value;render();});$('#m4search').addEventListener('input',e=>{q=e.target.value;render();});render();})();
"""

HTML_TMPL = """<!doctype html><html lang="ko" dir="ltr"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="color-scheme" content="light dark"><title>부품 4M 변경관리 현황</title>
<style>{CSS}</style></head><body>
<header class="top"><div class="bar">
<div class="brand"><b>부품 4M 변경관리 현황</b><span>Supplier Parts Change Management · 기준 {basis}</span></div>
<nav class="menu">
<button class="active" data-t="home">홈</button>
<button data-t="summary">Summary</button>
<button data-t="list">List</button>
<button data-t="m4">4M Issue 보고</button></nav>
<div class="dls">{DL}<button class="dl" id="dlhtml">⬇ HTML</button></div>
</div></header><main>

<section id="home" class="page show">
<h1 class="pt">진행중 현황 (홈)</h1>
<div class="sub">심의회·QA 인정 시험·GQMS 접수 단계와 정기/비정기 분해 중심 요약 · 4M 완료·Reject·4M 완료(C)·합계 열 제외</div>
<div class="kpis">{kpihtml}</div>
<div class="card"><h3><span class="b3"></span>정기/비정기 상태별 진행중 현황</h3>
<div class="sub" style="padding:8px 18px 0;margin:0">심의회·QA 인정 시험·GQMS 접수 합산 · 막대에 마우스를 올리면 수치가 표시됩니다</div>
<div class="pchart" id="pchart"></div></div>
<div class="card"><h3><span class="b3"></span>담당자별 진행중 구성 (심의회·QA·GQMS)</h3>
<div class="legend"><span><i style="background:var(--sim)"></i>심의회</span><span><i style="background:var(--qa)"></i>QA 인정 시험</span><span><i style="background:var(--gq)"></i>GQMS 접수</span></div>
<div class="chart" id="chart"></div></div>
<div class="card"><h3><span class="b3"></span>진행중 현황표 (정기/비정기별)</h3>{HOME_TBL}</div>
</section>

<section id="summary" class="page">
<h1 class="pt">Summary (전체)</h1>
<div class="sub">전체 진행 단계 × 정기/비정기 분해 요약 — 원본 Summary 시트 전체 열</div>
{FULL_TBL}
<div class="note">※ 심의회·QA 인정 시험·GQMS 접수 단계 아래 정기/비정기 값별 건수와 단계 '계'를 함께 표시합니다. 합계행 포함.</div>
</section>

<section id="list" class="page">
<h1 class="pt">List</h1><div class="sub">부품 4M 변경 상세 · 머리글 클릭 시 정렬 · 전체 컬럼은 원본 엑셀 다운로드</div>
<div class="toolbar"><div class="inp"><svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="11" cy="11" r="7"/><path d="M21 21l-4-4"/></svg>
<input id="lsearch" placeholder="전체 검색 (담당자, OEM, Part No, 변경점 …)"></div>
<select id="lper"><option value="25">25개씩</option><option value="50" selected>50개씩</option><option value="100">100개씩</option></select></div>
<div class="tablewrap listwrap"><table><thead id="lhead"></thead><tbody id="lbody"></tbody></table></div>
<div class="pager" id="lpager"></div></section>

<section id="m4" class="page">
<h1 class="pt">4M Issue 보고</h1><div class="sub">월례보고용 주요 4M 이슈 · <span id="m4count" class="count"></span></div>
<div class="toolbar"><div class="inp"><svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="11" cy="11" r="7"/><path d="M21 21l-4-4"/></svg>
<input id="m4search" placeholder="이슈 검색 (부품, 변경점, 협력사 …)"></div><select id="m4status"></select></div>
<div class="cards4m" id="m4cards"></div></section>

</main>
<footer>팀즈 백업 최신 파일 기준 · 매주 금요일 자동 갱신 · 홈(진행중)·Summary(전체)·List·4M Issue 보고 · 상단에서 HTML/원본 엑셀 다운로드</footer>
<script id="ldata" type="application/json">{LISTDATA}</script>
<script id="cdata" type="application/json">{CHART_JSON}</script>
<script id="pdata" type="application/json">{PERIOD_JSON}</script>
<script>{JS}</script></body></html>"""


def find_m4_sheet(wb):
    return find_sheet(wb, '4m', 'issue') or find_sheet(wb, '4m', '보고') \
        or find_sheet(wb, '4m') or find_sheet(wb, 'issue') or find_sheet(wb, '보고')

def copy_m4_sheet(dst_wb, src_ws, name='4M_Issue_보고'):
    """다른 워크북의 4M 시트(값)를 생성 엑셀로 복사해 자체 완결 + Risk 수식 작동을 보장."""
    if name in dst_wb.sheetnames:
        dst_wb.remove(dst_wb[name])
    ws = dst_wb.create_sheet(name)
    wrap = Alignment(wrap_text=True, vertical='top')
    for r in range(1, src_ws.max_row + 1):
        for c in range(1, src_ws.max_column + 1):
            v = src_ws.cell(row=r, column=c).value
            if v is not None:
                nc = ws.cell(row=r, column=c, value=v)
                nc.alignment = wrap
    return ws


# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description='부품 4M 변경관리 홈페이지 자동 생성')
    ap.add_argument('input', nargs='?', default=None,
                    help='List/Summary 입력 엑셀 (팀즈 999.백업 최신 파일). --download 시 생략 가능')
    ap.add_argument('--m4-file', default=None,
                    help='4M Issue 보고 시트를 가진 별도 엑셀 (예: VS협력사변경관리팀 업무 관리_v0.xlsx). '
                         '생략 시 input 안의 4M 시트를 사용')
    ap.add_argument('--download', action='store_true',
                    help='팀즈 공유 링크에서 최신 List 파일과 4M(업무관리) 파일을 직접 내려받아 사용 '
                         '(Microsoft Graph 토큰 필요: 환경변수 GRAPH_TOKEN)')
    ap.add_argument('--outdir', default='output')
    a = ap.parse_args()
    os.makedirs(a.outdir, exist_ok=True)

    # 인자 없이 실행한 경우: 토큰이 있으면 팀즈 다운로드, 없으면 주변 폴더의 원본 자동 사용
    if not a.input and not a.download:
        if _graph_token():
            a.download = True
        else:
            li, m4 = autofind_local([a.outdir])
            if li:
                a.input = li
                if m4 and not a.m4_file:
                    a.m4_file = m4
                print('입력 자동 선택(로컬 폴더에서 발견):')
                print('  · List :', a.input)
                print('  · 4M   :', a.m4_file or '(input 내 시트 사용)')
            else:
                print('실행할 입력을 찾지 못했습니다. 아래 중 하나로 실행하세요:')
                print('  1) 팀즈 자동 다운로드 : 환경변수 GRAPH_TOKEN 설정 후 다시 실행')
                print('                         (또는  python refresh_4m_homepage.py --download )')
                print('  2) 로컬 파일 지정     : python refresh_4m_homepage.py "<List.xlsx>" --m4-file "<업무관리.xlsx>"')
                print('  3) 같은 폴더에 원본 .xlsx 를 두고 다시 실행')
                print('     (예: Supplier Parts Change Management_v0_YYMMDD.xlsx, VS협력사변경관리팀 업무 관리_v0.xlsx)')
                sys.exit(1)

    # --download: 팀즈에서 최신 List/4M 파일을 직접 내려받아 input/m4_file 로 사용
    if a.download:
        try:
            a.input, a.m4_file = download_sources(a.outdir)
        except Exception as e:
            print('[다운로드 실패]', e)
            sys.exit(1)
        print('다운로드 완료:')
        print('  · List 최신 :', a.input)
        print('  · 4M 파일   :', a.m4_file)
    if not a.input:
        ap.error('입력 엑셀이 필요합니다. (--download 또는 입력 파일 경로)')

    basis = date_from_name(a.input)

    # 4M 시트 출처 결정: 별도 파일 우선
    m4_src_vals = None
    if a.m4_file:
        m4wb_vals = openpyxl.load_workbook(a.m4_file, data_only=True)
        m4_src_vals = find_m4_sheet(m4wb_vals)
        if m4_src_vals is None:
            print('경고: --m4-file 에서 4M Issue 시트를 찾지 못했습니다.')

    # 값 기준 분석 (List=input, 4M=별도파일 우선)
    wb = openpyxl.load_workbook(a.input, data_only=True)
    A = analyze(wb, m4ws=m4_src_vals)

    # 수식 포함 엑셀 생성
    wbf = openpyxl.load_workbook(a.input, data_only=False)
    if m4_src_vals is not None:                      # 4M 시트를 생성 엑셀에 복사(자체 완결)
        copy_m4_sheet(wbf, m4_src_vals, '4M_Issue_보고')
    A2 = analyze(openpyxl.load_workbook(a.input, data_only=True), m4ws=m4_src_vals)
    A2['listws'] = wbf['List'] if 'List' in wbf.sheetnames else find_sheet(wbf, 'list')
    A2['m4ws'] = find_m4_sheet(wbf)                  # 복사된(또는 기존) 시트
    build_summary_sheet(wbf, A2)
    xlsx_out = os.path.join(a.outdir, 'Supplier Parts Change Management_v0_'
                            + basis.replace('-', '')[2:] + '_Summary_정기비정기.xlsx')
    wbf.save(xlsx_out)
    try_recalc(xlsx_out)

    html_out = os.path.join(a.outdir, '부품4M_변경관리_현황.html')
    build_html(A, xlsx_out, html_out, basis)

    # 참고한 최신 원본 파일(List, 4M)을 outdir 에도 보존 → 함께 다운로드 가능
    import shutil
    saved_src = []
    for src in [a.input, a.m4_file]:
        if not src:
            continue
        dst = os.path.join(a.outdir, os.path.basename(src))
        if os.path.abspath(src) != os.path.abspath(dst):
            try:
                shutil.copy2(src, dst); saved_src.append(dst)
            except Exception as e:
                print('원본 보존 실패:', src, e)
        else:
            saved_src.append(dst)

    print('기준일      :', basis)
    print('원본 보존   :', [os.path.basename(p) for p in saved_src])
    print('4M 출처     :', a.m4_file or '(input 내 시트)')
    print('담당자      :', len(A['people']), '명', A['people'])
    print('진행 단계   :', A['stages'])
    print('정기/비정기 :', A['periods'])
    print('4M Issue    :', sum(1 for r in range((A['m4ws'].max_row if A['m4ws'] else 0)))
          if A['m4ws'] else 0, '행 스캔')
    print('엑셀 생성   :', xlsx_out)
    print('HTML 생성   :', html_out)


if __name__ == '__main__':
    main()
